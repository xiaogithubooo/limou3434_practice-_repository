# from openpyxl import load_workbook
# wb = load_workbook(r"C:\Users\Limou_p350ml9\Desktop\Test\新建文件夹\pythontest\workbook1.xlsx")
# wb.template = True
# wb.save(r'C:\Users\Limou_p350ml9\Desktop\Test\新建文件夹\pythontest\document_template.xltx')

from openpyxl import *

# 加载模板文件
template_wb = load_workbook(r'C:\Users\Limou_p350ml9\Desktop\Test\新建文件夹\pythontest\document_template.xltx')

# 创建一个新的工作簿
new_wb = Workbook()

# 复制模板的样式到新工作簿
for sheet in template_wb.sheetnames:
    new_sheet = new_wb.create_sheet(title=sheet)
    for row in template_wb[sheet].iter_rows(min_row=1, max_row=template_wb[sheet].max_row, max_col=template_wb[sheet].max_column):
        new_row = [cell.value for cell in row]
        new_sheet.append(new_row)

# 保存新工作簿
new_wb.save(r'C:\Users\Limou_p350ml9\Desktop\Test\新建文件夹\pythontest\new_document.xlsx')