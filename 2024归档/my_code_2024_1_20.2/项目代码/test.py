'''
本代码只是暂时供参考使用，没有使用过多的技巧
'''

import matplotlib.pyplot as plt # 可视化绘图库
from openpyxl import * # 操作 excel 的库
import pathlib # 路径库（后期将路径修改为 Path 对象，可以具备一定的可移植性）
from re import * # 正则表达式库

'''
一、管理员
使用一些 UI 界面，引导管理员设置课程标准以及相关的字段信息
TUDO...
'''

'''
二、使用者
使用可视化界面，根据管理员的标准，简化使用者填入数据的过程...
TUDO...
'''

'''
三、原始数据
假设此时已经得到数据（本例体现为表格，实际操作可能是一个经过处理的 JSON 文件）
TUDO...
'''

'''
四、程序处理
1.初步处理数据：结合管理员和使用者提供的数据，初始待处理的数据
后期可以用 JSON 存放数据，这里使用 excel 存放数据，并且暴力遍历分析有用的信息字段，以及相关的信息统计...
'''

# 存储信息字段的变量
filename = '.\项目代码\面向对象程序设计课程-目标达成度分析.xlsx'
workbook = load_workbook(filename) # 打开表格文件
worksheet = workbook['Sheet2'] # 获取工作表
merged_cells = worksheet.merged_cells.ranges # 获取所有被合并的单元格

# 遍历合并单元格获取信息字段
usual_performance = 0 # 平时成绩占比
final_grade = 0 # 期末成绩占比
Curriculum_objective_count = 0 # 课程目标数量
questions_count = 0 # 题目数量

for merged_range in worksheet.merged_cells.ranges:
    # 获取合并单元格的起始行、起始列
    start_row, start_col = merged_range.min_row, merged_range.min_col
    # 获取合并单元格的值（即范围内的第一个单元格的值）
    cell_value = worksheet.cell(row=start_row, column=start_col).value
    # 打印合并单元格的内容
    # print(f"{start_row}, {start_col}: {cell_value}")

    if '平时成绩' in cell_value:
        match = search(r'\b(\w+)\%', cell_value)
        if match:
            usual_performance = match.group()
    elif '期末成绩' in cell_value:
        match = search(r'\b(\w+)\%', cell_value)
        if match:
            final_grade = match.group()
    elif ('课程目标' in cell_value) and (len(cell_value) < 10):
        match = search(r'课程目标(\d)', cell_value)
        if match:
            Curriculum_objective_count += 1

# 遍历单元格获取全部字段
for row in worksheet.iter_rows():
    for cell in row:
        if cell.value is not None:
            match = search(r'题目\d', str(cell.value))
            if match:
                # print(f"单元格 {cell.coordinate} 中存在 '题目number' 内容")
                questions_count += 1

print('平时成绩占比:' + usual_performance)
print('期末成绩占比:' + final_grade)
print(f'课程目标数量:{Curriculum_objective_count}')
print(f'题目数量:{questions_count}')

workbook.close()

# 假设这里通过各个字段信息生成了各种数据
# TUDO...

"""
2.生成而可视化图形
  2.1.得到每个课程目标的达成度
  2.2.得到每个课程目标的达成度分布图
  ...
"""

# 假设得到了课程达成度数据
# 获取工作表对象
workbook = load_workbook(filename, data_only=True) # 打开表格文件
worksheet = workbook['Sheet2'] # 获取工作表

program_objective_1 = worksheet['E41'].value
program_objective_2 = worksheet['G41'].value
program_objective_3 = worksheet['I41'].value
program_objective_4 = worksheet['K41'].value

print(program_objective_1)
print(program_objective_2)
print(program_objective_3)
print(program_objective_4)

datas = [program_objective_1, program_objective_2, program_objective_3, program_objective_4]

# 创建画布和坐标轴对象
fig, ax = plt.subplots()

# 绘制矩形统计图
bar_rects = ax.bar(range(len(datas)), datas)

# 设置坐标轴标签
ax.set_xlabel(f'课程目标')
ax.set_ylabel(f'课程目标达成度')

# 设置字体
plt.rcParams['font.sans-serif'] = ['SimHei']

# 标注每一根矩形柱
i = 1
for rect in bar_rects:
    height = rect.get_height()
    ax.text(rect.get_x() + rect.get_width() / 2, height, f'课程目标{i}的达成度\n' + str(height),
            ha='center', va='bottom')
    i += 1

# 显示图形
plt.show()